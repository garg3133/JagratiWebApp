# standard library
import json
from datetime import datetime, date

# third-party
from openpyxl import load_workbook

# Django
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.sites.shortcuts import get_current_site
from django.core import serializers
from django.core.mail import send_mail
# from django.core.serializers.json import DjangoJSONEncoder
from django.forms.models import model_to_dict
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.template.defaulttags import register
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.encoding import force_bytes
from django.utils.html import strip_tags
from django.utils.http import urlsafe_base64_encode

# local Django
from accounts.tokens import account_activation_token
from .models import(
    Calendar, ClassworkHomework, Feedback, Schedule, Section,
    Student, StudentAttendence, StudentSchedule, Volunteer,
    VolunteerAttendence, VolunteerSchedule, UpdateScheduleRequest,
)


# Create your views here.
def index(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'home/index.html')

@login_required
def completeProfile(request):
    """ For completing the Profile after successful signup and activation of account.
        Mandatory before accessing the Dashboard."""

    # Redirect to Dashboard if Profile is already complete
    if Volunteer.objects.filter(email=request.user).exists():
        return redirect('dashboard')

    if request.method == 'POST':
        roll_no         = request.POST['roll_no']
        first_name      = request.POST['first_name']
        last_name       = request.POST['last_name']
        gender          = request.POST['gender']
        alt_email       = request.POST['alt_email']
        batch           = request.POST['batch']
        programme       = request.POST['programme']
        street_address1 = request.POST['street_address1']
        street_address2 = request.POST['street_address2']
        pincode         = request.POST['pincode']
        city            = request.POST['city']
        state           = request.POST['state']
        dob             = request.POST['dob']
        contact_no      = request.POST['contact_no']
        profile_image 	= request.FILES.get('profile_image')

        volun = Volunteer(
            email           = request.user,
            roll_no         = roll_no,
            first_name      = first_name,
            last_name       = last_name,
            profile_image	= profile_image,
            gender          = gender,
            city            = city,
            state           = state,
            dob             = dob,
            contact_no      = contact_no,
            pincode         = pincode,
            programme       = programme,
            batch           = batch,
            alt_email       = alt_email,
            street_address1 = street_address1,
            street_address2 = street_address2,
        )
        volun.save()

        # Notify Admin for New User Sign Up
        current_site = get_current_site(request)

        from_email = settings.DEFAULT_FROM_EMAIL
        to = settings.ADMINS_EMAIL
        subject = '[noreply] New User Signed Up'
        html_message = render_to_string('accounts/email/account_authentication_email.html', {
            'volun': volun,
            'domain': current_site.domain,
            'uid':urlsafe_base64_encode(force_bytes(volun.email.pk)),
            'token':account_activation_token.make_token(volun.email),
        })
        plain_message = strip_tags(html_message)
        send_mail(
            subject, plain_message, from_email, to,
            fail_silently=False, html_message=html_message,
        )

        logout(request)
        return redirect('profile_completed')
    
    context = {
        'last_5_year': datetime.now().year - 5,   # For Batch Dropdown
    }
    return render(request, 'home/set_profile.html', context)

# @register.filter
# def get_item(choices, key):
# 	dictionary = {key: value for key, value in choices}
# 	return dictionary.get(key)
# https://docs.djangoproject.com/en/2.2/howto/custom-template-tags/#writing-custom-template-filters

# DON'T TOUCH THE DASHBOARD
@login_required
def dashboard(request):
    volun = Volunteer.objects.filter(email=request.user)
    if volun.exists():
        volun = volun[0]
    else:
        return redirect('set_profile')

    # Update today's date in Calendar if not already there
    # MUST BE REMOVED AFTER CALENDAR IS CREATED
    today_cal = Calendar.objects.filter(date=date.today())
    if today_cal.exists():
        today_cal = today_cal[0]
    else:
        today_cal_new = Calendar(date = date.today())
        today_cal_new.save()
        today_cal = Calendar.objects.get(date=date.today())


    if request.method == 'POST':
        if request.POST.get('submit') == 'class-info':
            class_info_date_str		= request.POST['class-info-date']
            class_info_section		= request.POST['class-info-section']
            
            class_info_date = datetime.strptime(class_info_date_str, '%Y-%m-%d').date()
            class_info_day = class_info_date.strftime("%A")

            calendar = Calendar.objects.filter(date=class_info_date)

            # If calendar instance for that day is not created
            if not calendar.exists():
                context = {
                    #dash-main
                    'no_class_found' : "bilkul_nhi",
                    'selected_date' : class_info_date,
                    'selected_schedule' : class_info_section,
                }
                # context.update(context1)

                return render(request, 'home/dashboard.html', context)

            calendar = calendar[0]

            # If No Class is Scheduled on that day
            if calendar.class_scheduled is False:
                context = {
                    #dash-main
                    'no_class_scheduled' : "haan",
                    'calendar_remark' : calendar.remark,
                    'selected_date' : class_info_date,
                    'selected_schedule' : class_info_section,
                }
                # context.update(context1)

                return render(request, 'home/dashboard.html', context)

            section = Section.objects.get(section_id=class_info_section)
            schedule = Schedule.objects.filter(day=class_info_day, section=section)
            if schedule.exists():
                if class_info_date > date.today():
                    students_attended = "Class not yet scheduled!"
                    vol_volunteered = "Class not yet scheduled!"
                else:
                    students_attended = StudentAttendence.objects.filter(date = calendar, present = True).order_by('sid__school_class')
                    vol_volunteered = VolunteerAttendence.objects.filter(date = calendar, present = True).order_by('roll_no__roll_no')

                    # USE A LIST INSTEAD [G, A, C, M, S]
                    # Students present from each village
                    stu_att_g = 0
                    stu_att_a = 0
                    stu_att_c = 0
                    stu_att_m = 0
                    stu_att_s = 0
                    stu_att_ms = 0
                    if students_attended.exists():
                        for stu_att in students_attended:
                            stu_vill = stu_att.sid.village
                            if stu_vill == 'G':
                                stu_att_g += 1
                            elif stu_vill == 'A':
                                stu_att_a += 1
                                stu_att_ms += 1
                            elif stu_vill == 'C':
                                stu_att_c += 1
                                stu_att_ms += 1
                            elif stu_vill == 'M':
                                stu_att_m += 1
                                stu_att_ms += 1
                            elif stu_vill == 'S':
                                stu_att_s += 1
                                stu_att_ms += 1

                    stu_att_village = {
                        'stu_att_g': stu_att_g,
                        'stu_att_a': stu_att_a,
                        'stu_att_c': stu_att_c,
                        'stu_att_m': stu_att_m,
                        'stu_att_s': stu_att_s,
                        'stu_att_ms': stu_att_ms,
                    }

                    if class_info_date == date.today() and not students_attended.exists():
                        students_attended = "Not yet updated!"
                    elif not students_attended.exists():
                        students_attended = "No student present."

                    if class_info_date == date.today() and not vol_volunteered.exists():
                        vol_volunteered = "Not yet updated!"
                    elif not vol_volunteered.exists():
                        vol_volunteered = "No volunteers were present."

                # Classwork-Homework starts
                cw_hw = ClassworkHomework.objects.filter(date=class_info_date, section=section)
                if cw_hw.exists():
                    context = {
                        #dash-main
                        # 'schedule' : schedule,
                        'students_attended' : students_attended,
                        'stu_att_village' : stu_att_village,
                        'vol_volunteered' : vol_volunteered,
                        'cw_hw' : cw_hw[0],
                        'selected_date' : class_info_date,
                        'selected_schedule' : class_info_section,
                    }
                    # context.update(context1)

                    return render(request, 'home/dashboard.html', context)
                else:
                    context = {
                        #dash-main
                        # 'schedule' : schedule,
                        'students_attended' : students_attended,
                        'stu_att_village' : stu_att_village,
                        'vol_volunteered' : vol_volunteered,
                        'cw_hw' : {
                            'cw' : "Not yet updated!",
                            'hw' : "Not yet updated!",
                        },
                        'selected_date' : class_info_date,
                        'selected_schedule' : class_info_section,
                    }
                    # context.update(context1)

                    return render(request, 'home/dashboard.html', context)
            else:
                # Keeping it just is case if Ajax takes time to load. DISABLE SUBMIT BUTTON UNTIL AJAX LOADS.
                context = {
                    #dash-main
                    'no_schedule_found' : "yup!",
                    'selected_date' : class_info_date,
                    'selected_schedule' : class_info_section,
                }
                # context.update(context1)

                return render(request, 'home/dashboard.html', context)  # The chosen section is not taught on the chosen day

        # TRY UPDATING CW_HW WITH AJAX  --- NO! MAY RESULT IN DUPLICATE REQUESTS
        elif request.POST.get('submit') == 'update-cwhw':
            """ This POST Request won't be Generated (because there will be no means to generated it in the template) if
                1. Selected Date is not present in Calendar. (Already checked for above)
                2. Selected Day (from Date) and Section not present in Schedule. (Already checked for above and
                    submit button will be disabled until AJAX is loaded completely)
                3. No Class Scheduled on Selected Date. (Already checked for above)
            """
            # Date and section selected before pressing submit button
            cwhw_date_str			= request.POST['date']
            cwhw_section			= request.POST['section']
            cw						= request.POST['cw']
            hw						= request.POST['hw']
            comment					= request.POST['comment']

            cwhw_date = datetime.strptime(cwhw_date_str, '%Y-%m-%d').date()

            cal_date = Calendar.objects.get(date = cwhw_date)
            section = Section.objects.get(section_id=cwhw_section)

            cw_hw = ClassworkHomework.objects.filter(date=cal_date, section=section)
            if cw_hw.exists():
                cw_hw = cw_hw[0]
                if cw:
                    cw_hw.cw += '\n' + cw + '\n - ' + volun.first_name + ' ' + volun.last_name + ', ' + volun.roll_no + '\n'
                if hw:
                    cw_hw.hw += '\n' + hw + '\n - ' + volun.first_name + ' ' + volun.last_name + ', ' + volun.roll_no + '\n'
                if comment:
                    cw_hw.comment += '\n' + comment + '\n - ' + volun.first_name + ' ' + volun.last_name + ', ' + volun.roll_no + '\n'
                cw_hw.save()
            else:
                if cw:
                    cw += '\n - ' + volun.first_name + ' ' + volun.last_name + ', ' + volun.roll_no + '\n'
                if hw:
                    hw += '\n - ' + volun.first_name + ' ' + volun.last_name + ', ' + volun.roll_no + '\n'
                if comment:
                    comment += '\n - ' + volun.first_name + ' ' + volun.last_name + ', ' + volun.roll_no + '\n'
                cw_hw = ClassworkHomework(date=cal_date, section=section, cw=cw, hw=hw, comment=comment)
                cw_hw.save()

            # CONTEXT IS NOT BEING USED CURRENTLY
            context = {
                #dash-main
                'class_info_submitted' : "nopes",

                # CAN USE SESSION VARIABLES TO SEND THESE DATA
                # OR NOT THAT NECESSARY
                'selected_date' : cwhw_date,   # <-- New
                'selected_schedule' : cwhw_section,   # <-- New
            }
            # context.update(context1)

            messages.success(request, 'CW_HW update successful!')
            # return render(request, 'home/dashboard.html', context)
            return HttpResponseRedirect(reverse('dashboard'))

    else:
        context = {
            #dash-main
            'class_info_submitted' : "nopes",
        }
        # context.update(context1)

        return render(request, 'home/dashboard.html', context)

def showSectionInDashboard(request):
    class_info_date_str = request.GET.get('class_date', None)
    class_info_date = datetime.strptime(class_info_date_str, '%Y-%m-%d').date()
    class_info_day = class_info_date.strftime("%A")

    schedule = Schedule.objects.filter(day=class_info_day).order_by('section__section_id')

    data = {}

    for sch in schedule:
        data[sch.section.section_id] = sch.section.name

    return JsonResponse(data)

@login_required
def updateSchedule(request):
    volun = Volunteer.objects.filter(email=request.user)
    if volun.exists():
        volun = volun[0]
    else:
        return redirect('set_profile')

    context = {
        #dash-schedule
        'day': Schedule.DAY,
        'update_req' : UpdateScheduleRequest.objects.filter(volunteer=volun).order_by('-date'),
        'last_update_req' : UpdateScheduleRequest.objects.filter(volunteer=volun, approved=False, declined=False, by_admin=False, cancelled=False),
    }
    
    if request.method == 'POST':
        if request.POST.get('submit') == 'update-schedule':
            new_day		= request.POST['day']
            new_section	= request.POST['section']

            section = Section.objects.get(section_id=new_section)
            schedule = Schedule.objects.get(day=new_day, section=section)

            prev_update_req = UpdateScheduleRequest.objects.filter(volunteer=volun, approved=False, declined=False, by_admin=False, cancelled=False)
            if prev_update_req.exists():
                prev_update_req = prev_update_req[0]
                prev_update_req.cancelled = True
                # For cpanel
                # prev_update_req.approved = False
                # prev_update_req.declined = False
                # prev_update_req.by_admin = False
                prev_update_req.save()

            prev_vol_sch = VolunteerSchedule.objects.filter(roll_no=volun)
            if prev_vol_sch.exists():
                prev_sch = prev_vol_sch[0].schedule
                update_req = UpdateScheduleRequest(volunteer=volun, previous_schedule=prev_sch, updated_schedule=schedule)
                update_req.save()
            else:
                update_req = UpdateScheduleRequest(volunteer=volun, updated_schedule=schedule)
                update_req.save()

            messages.success(request, 'Schedule update requested successfully!')
            return HttpResponseRedirect(reverse('update_schedule'))

        elif request.POST.get('submit') == 'cancel-last-req':
            prev_update_req = UpdateScheduleRequest.objects.filter(volunteer=volun, approved=False, declined=False, by_admin=False, cancelled=False)
            if prev_update_req.exists():
                prev_update_req = prev_update_req[0]
                prev_update_req.cancelled = True
                # For cpanel
                # prev_update_req.approved = False
                # prev_update_req.declined = False
                # prev_update_req.by_admin = False
                prev_update_req.save()

            messages.success(request, 'Last request cancelled successfully!')
            return HttpResponseRedirect(reverse('update_schedule'))

    return render(request, 'home/update_schedule.html', context)

def showSectionInUpdateSchedule(request):
    sch_day = request.GET.get('sch_day', None)

    schedule = Schedule.objects.filter(day=sch_day).order_by('section__section_id')

    data = {}

    for sch in schedule:
        data[sch.section.section_id] = sch.section.name

    return JsonResponse(data)

@login_required
def updateProfile(request):
    volun = Volunteer.objects.filter(email=request.user)
    if volun.exists():
        volun = volun[0]
    else:
        return redirect('set_profile')

    context = {
        #dash-update
        'last_5_year': datetime.now().year - 5,
    }
    
    if request.method == 'POST':
        roll_no         = request.POST['roll_no']
        first_name      = request.POST['first_name']
        last_name       = request.POST['last_name']
        gender          = request.POST['gender']
        alt_email       = request.POST['alt_email']
        batch           = request.POST['batch']
        programme       = request.POST['programme']
        street_address1 = request.POST['street_address1']
        street_address2 = request.POST['street_address2']
        pincode         = request.POST['pincode']
        city            = request.POST['city']
        state           = request.POST['state']
        dob             = request.POST['dob']
        contact_no      = request.POST['contact_no']
        profile_image	= request.FILES.get('profile_image')

        if roll_no:
            if volun.roll_no != roll_no:
                duplicate_roll_check = Volunteer.objects.filter(roll_no=roll_no)
                if duplicate_roll_check.exists():
                    context1 = {
                        'update_error': "A volunteer with entered roll no. already exists."
                    }
                    context.update(context1)
                    messages.error(request, 'Profile update failed!')
                    return render(request, 'home/update_profile.html', context)
                else:
                    volun.roll_no = roll_no
        if first_name:
            volun.first_name = first_name
        if last_name:
            volun.last_name = last_name
        volun.gender = gender
        volun.batch = batch
        volun.programme = programme
        volun.dob = dob
        if contact_no:
            volun.contact_no = contact_no
        if alt_email:
            volun.alt_email = alt_email
        if street_address1:
            volun.street_address1 = street_address1
        if street_address2:
            volun.street_address2 = street_address2
        if city:
            volun.city = city
        if state:
            volun.state = state
        if pincode:
            volun.pincode = pincode
        if 'profile_image' in request.FILES:
            # delete the previous one 
            volun.profile_image.delete(False)
            volun.profile_image = profile_image

        volun.save()

        messages.success(request, 'Profile updated Successfully!')
        return HttpResponseRedirect(reverse('update_profile'))

    return render(request, 'home/update_profile.html', context)

@login_required
def studentsAttendence(request):
    volun = Volunteer.objects.filter(email=request.user)
    if volun.exists():
        volun = volun[0]
    else:
        return redirect('set_profile')

    today_cal = Calendar.objects.filter(date=date.today())
    # Update today's date in Calendar if not already there
    if today_cal.exists():
        today_cal = today_cal[0]
    else:
        today_cal_new = Calendar(date = date.today())
        today_cal_new.save()
        today_cal = Calendar.objects.get(date=date.today())

    # For Creating Empty Student Attendence Instances
    if today_cal.class_scheduled is True and not StudentAttendence.objects.filter(date__date = date.today()).exists():
        today_stu_sch = StudentSchedule.objects.filter(day=date.today().strftime("%A"))
        for stu_sch in today_stu_sch:
            stu_attendance = StudentAttendence(sid = stu_sch.sid, date = today_cal)
            stu_attendance.save()

    if today_cal.class_scheduled is True and StudentAttendence.objects.filter(date__date = date.today()).count() != StudentSchedule.objects.filter(day=date.today().strftime("%A")).count():
        today_stu_sch = StudentSchedule.objects.filter(day=date.today().strftime("%A"))
        for stu_sch in today_stu_sch:
            if not StudentAttendence.objects.filter(sid = stu_sch.sid, date = today_cal).exists():
                stu_attendance = StudentAttendence(sid = stu_sch.sid, date = today_cal)
                stu_attendance.save()

    # If no Class is Scheduled
    no_class_today = ''
    if today_cal.class_scheduled is False:
        no_class_today = 'yesss!'

    context = {
        # Overall
        'no_class_today' : no_class_today,

        #dash-stu-att
        'today_date' : date.today(),
        'today_stu' : StudentAttendence.objects.filter(date = today_cal, sid__school_class__range = (1, 3)).order_by('sid__school_class', 'sid__first_name', 'sid__last_name'),
    }

    if request.method == 'POST':
        today_date = Calendar.objects.get(date=date.today())
        stu_array = request.POST.getlist('attended')
        selected_class = request.POST['selected_class']

        class_range = selected_class.split('-')
        class_range_min = class_range[0]
        class_range_max = class_range[1]

        # Mark everyone's absent
        stu_today = StudentAttendence.objects.filter(date = today_date, sid__school_class__range = (class_range_min, class_range_max))
        for stu in stu_today:
            stu.present = False
            stu.hw_done = False
            stu.save()

        for sid in stu_array:
            stu = Student.objects.get(id=sid)
            stu_attendance = StudentAttendence.objects.filter(sid = stu, date = today_date)[0]
            stu_attendance.present = True
            stu_attendance.hw_done = False
            stu_attendance.save()

        messages.success(request, 'Attendence marked successfully!')
        return HttpResponseRedirect(reverse('stu_attendence'))

    return render(request, 'home/stu_attendence.html', context)

def studentAttendenceAjax(request):
    stu_class = request.GET.get('stu_class', None)

    class_range = stu_class.split('-')
    class_range_min = class_range[0]
    class_range_max = class_range[1]

    today_cal = Calendar.objects.get(date=date.today())

    stu_att_to_show = StudentAttendence.objects.filter(
        date=today_cal,
        sid__school_class__range=(class_range_min, class_range_max),
    ).order_by('sid__school_class', 'sid__first_name', 'sid__last_name')
    
    data = {}

    for stu_att in stu_att_to_show:
        key = str(stu_att.sid.school_class) + stu_att.sid.first_name + stu_att.sid.last_name  # For sorting purpose.
        data[key] = [stu_att.sid.id, stu_att.sid.first_name, stu_att.sid.last_name, stu_att.sid.school_class, stu_att.present]
    
    # json_data = json.dumps(data)
    # json_stu_to_show = serializers.serialize("json", stu_to_show)
    return JsonResponse(data)

@login_required
def volunteersAttendence(request):
    volun = Volunteer.objects.filter(email=request.user)
    if volun.exists():
        volun = volun[0]
    else:
        return redirect('set_profile')

    today_cal = Calendar.objects.filter(date=date.today())
    # Update today's date in Calendar if not already there
    if today_cal.exists():
        today_cal = today_cal[0]
    else:
        today_cal_new = Calendar(date = date.today())
        today_cal_new.save()
        today_cal = Calendar.objects.get(date=date.today())

    # For Creating Empty Volunteer Attendence Instances
    if today_cal.class_scheduled is True and not VolunteerAttendence.objects.filter(date__date=date.today()).exists():
        today_vol_sch = VolunteerSchedule.objects.filter(day=date.today().strftime("%A"))
        for vol_sch in today_vol_sch:
            vol_attendance = VolunteerAttendence(roll_no=vol_sch.roll_no, date=today_cal)
            vol_attendance.save()
    
    # If no Class is Scheduled
    no_class_today = ''
    if not today_cal.class_scheduled:
        no_class_today = 'yesss!'

    context = {
        # Overall
        'no_class_today' : no_class_today,

        #dash-vol-att
        'today_date' : date.today(),
        'today_volun' : VolunteerAttendence.objects.filter(date=today_cal).order_by('roll_no__roll_no'),
    }
    
    if request.method == 'POST':
        today_date = Calendar.objects.get(date=date.today())
        vol_array = request.POST.getlist('volunteered')
        extra_vol_array = request.POST.getlist('extra-vol')

        # Mark everyone's absent
        vol_today = VolunteerAttendence.objects.filter(date=today_date)
        for vol in vol_today:
            vol.present = False

            # For cpanel
            if vol.extra is None:
                vol.extra = False

            vol.save()

        for vol in vol_array:
            roll_no = Volunteer.objects.get(id=vol) #i is first column of volun_array
            vol_attendance = VolunteerAttendence.objects.get(roll_no=roll_no, date=today_date)
            vol_attendance.present = True

            # For cpanel
            if vol_attendance.extra is None:
                vol_attendance.extra = False

            vol_attendance.save()

        for extra_vol in extra_vol_array:
            roll_no = Volunteer.objects.filter(roll_no=extra_vol)
            if roll_no.exists():
                extra_vol_att = VolunteerAttendence.objects.filter(roll_no=roll_no[0], date=today_date)
                if not extra_vol_att.exists():
                    extra_vol_att = VolunteerAttendence(roll_no=roll_no[0], date=today_date, present=True, extra=True)
                    extra_vol_att.save()

        messages.success(request, 'Attendence marked successfully!')
        return HttpResponseRedirect(reverse('vol_attendence'))

    return render(request, 'home/vol_attendence.html', context)

@login_required
def volunteersList(request):
    context = {
        'day': Schedule.DAY,
    }
    return render(request, 'home/vol_list.html', context)

def volunteerListAjax(request):
    vol_list_day = request.GET.get('vol_list_day', None)

    data = {}

    if not vol_list_day:
        return JsonResponse(data)

    vol_to_show = VolunteerSchedule.objects.filter(day = vol_list_day)

    # setion_id used as key to display volunteers sorted by section_id and roll_no to make every key unique
    for vol_sch in vol_to_show:
        key = str(vol_sch.schedule.section.section_id) + str(vol_sch.roll_no.roll_no)
        data[key] = [vol_sch.roll_no.roll_no, vol_sch.roll_no.first_name, vol_sch.roll_no.last_name, vol_sch.schedule.section.name]


    # json_data = json.dumps(data, cls=DjangoJSONEncoder)
    # json_vol_list = serializers.serialize("json", vol_list)
    return JsonResponse(data)

def feedback(request):
    submitted = ''
    if request.method == 'POST':
        anonymous_array	= request.POST.getlist('anonymousCheck')
        name			= request.POST['name']
        roll_no			= request.POST['rollNo']
        email			= request.POST['email']
        feedback		= request.POST['feedback']
        if len(anonymous_array) != 0:
            feedback = Feedback(name = 'Anonymous', feedback = feedback)
            feedback.save()
        else:
            feedback = Feedback(name=name, roll_no=roll_no, email=email, feedback=feedback)
            feedback.save()

        submitted = "yesssss!"
    context = {
        'logout_redirect_site' : 'feedback',
        'submitted' : submitted,
    }
    return render(request, 'home/feedback.html', context)

def update_students(request):
    path = "D:/Python Projects/JagratiWebApp/Jagrati/home/student.xlsx"

    wb_obj = load_workbook(path)

    sheet_obj = wb_obj.active 

    max_row = sheet_obj.max_row

    for i in range(3, max_row+1):
        first_name = sheet_obj.cell(row = i, column = 2).value
        last_name = sheet_obj.cell(row = i, column = 3).value
        school_class = sheet_obj.cell(row = i, column = 4).value
        village = sheet_obj.cell(row = i, column = 5).value
        guardian_name = sheet_obj.cell(row = i, column = 6).value
        contact_no = sheet_obj.cell(row = i, column = 7).value

        if not Student.objects.filter(first_name = first_name, last_name = last_name, school_class = school_class, village = village, guardian_name = guardian_name).exists():
            if contact_no is None:
                student = Student(first_name = first_name, last_name = last_name, school_class = school_class, village = village, guardian_name = guardian_name)
                student.save()
            else:
                student = Student(first_name = first_name, last_name = last_name, school_class = school_class, village = village, guardian_name = guardian_name, contact_no = int(contact_no))
                student.save()
            for day, day2 in Schedule.DAY:
                print(day)
                stu_sch = StudentSchedule(sid = student, schedule = Schedule.objects.get(day = day, section__section_id = '4A'))
                stu_sch.save()
    
    return HttpResponse('Updated successfully!')