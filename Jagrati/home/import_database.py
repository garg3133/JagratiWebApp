from openpyxl import load_workbook
# from home.models import Student

path = "/home/priyansh/Python Projects/JagratiWebApp/Jagrati/home/student.xlsx"

wb_obj = load_workbook(path)

sheet_obj = wb_obj.active 

max_row = sheet_obj.max_row

for i in range(3, 8):
    first_name = sheet_obj.cell(row = i, column = 2).value
    last_name = sheet_obj.cell(row = i, column = 3).value
    school_class = sheet_obj.cell(row = i, column = 4).value
    village = sheet_obj.cell(row = i, column = 5).value
    guardian_name = sheet_obj.cell(row = i, column = 6).value
    contact_no = sheet_obj.cell(row = i, column = 7).value
    
    if contact_no is not None:
        print(int(contact_no))

    # if not Student.objects.filter(first_name = first_name, last_name = last_name, school_class = school_class, village = village, guardian_name = guardian_name).exists()
    #     if contact_no is None:
    #         student = Student(first_name = first_name, last_name = last_name, school_class = school_class, village = village, guardian_name = guardian_name)
    #         student.save()
    #     else:
    #         student = Student(first_name = first_name, last_name = last_name, school_class = school_class, village = village, guardian_name = guardian_name, contact_no = contact_no)
    #         student.save()
    
