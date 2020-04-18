from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import(
	Volunteer,
	Student,
	Schedule,
	VolunteerSchedule,
	StudentSchedule,
	ClassworkHomework,
	Calendar,
	VolunteerAttendence,
	StudentAttendence,
	Feedback,
	UpdateScheduleRequest,
)
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from datetime import datetime, date
from django.core import serializers
from django.template.defaulttags import register
import json
# from django.core.serializers.json import DjangoJSONEncoder
from django.forms.models import model_to_dict
from openpyxl import load_workbook
from django.contrib import messages
from django.urls import reverse

# Create your views here.
def index(request):
	if request.user.is_authenticated:
		return redirect('dashboard')
	return render(request, 'home/index.html')

@login_required
def completeProfile(request):
	""" For completing the Profile after successful signup and activation of account.
		Mandatory before accessing the Dashboard."""

	# Redirect to Dashboard if Profile is already complete
	if Volunteer.objects.filter(email=request.user).exists():
		return redirect('dashboard')

	if request.method == 'POST':
		roll_no         = request.POST['roll_no']
		first_name      = request.POST['first_name']
		last_name       = request.POST['last_name']
		gender          = request.POST['gender']
		alt_email       = request.POST['alt_email']
		batch           = request.POST['batch']
		programme       = request.POST['programme']
		street_address1 = request.POST['street_address1']
		street_address2 = request.POST['street_address2']
		pincode         = request.POST['pincode']
		city            = request.POST['city']
		state           = request.POST['state']
		dob             = request.POST['dob']
		contact_no      = request.POST['contact_no']
		vol_obj = Volunteer(
			email           = email,
			roll_no         = roll_no,
			first_name      = first_name,
			last_name       = last_name,
			gender          = gender,
			city            = city,
			state           = state,
			dob             = dob,
			contact_no      = contact_no,
			pincode         = pincode,
			programme       = programme,
			batch           = batch,
			alt_email       = alt_email,     
			street_address1 = street_address1,
			street_address2 = street_address2,
		)
		vol_obj.save()

		messages.success(request, 'Profile Saved Successfully!')
		return HttpResponseRedirect(reverse('dashboard'))
	
	context = {
		'last_4_year': datetime.now().year - 4,   # For Batch Dropdown
	}
	return render(request, 'home/set_profile.html', context)

# @register.filter
# def get_item(choices, key):
# 	dictionary = {key: value for key, value in choices}
# 	return dictionary.get(key)
# https://docs.djangoproject.com/en/2.2/howto/custom-template-tags/#writing-custom-template-filters

# DON'T TOUCH THE DASHBOARD
def dashboard(request):
	if request.user.is_authenticated:
		volun = Volunteer.objects.filter(email=request.user)
		if volun.exists():
			volun = volun[0]
		else:
			return redirect('set_profile')

		# Update today's date in Calendar if not already there
		# MUST BE REMOVED AFTER CALENDAR IS CREATED
		today_cal = Calendar.objects.filter(date=date.today())
		if today_cal.exists():
			today_cal = today_cal[0]
		else:
			today_cal_new = Calendar(date = date.today())
			today_cal_new.save()
			today_cal = Calendar.objects.get(date=date.today())

		# For main dash ajax section display names
		# choices_dict = {key: value for key, value in Schedule.SECTION}
		# choices_dict_json = json.dumps(choices_dict)

		# DASHBOARD HAVE NOTHING WHICH CAN ONLY BE DONE ON THE PRESENT DAY
		# If no Class is Scheduled
		# no_class_today = ''
		# if not today_cal.class_scheduled:
		# 	no_class_today = 'yesss!'

		context = {
			# Overall
			# 'no_class_today' : no_class_today,

			#dash-main
			'choices': Schedule.SECTION,
			# 'choices_dict': choices_dict_json,
		}
		
		if request.method == 'POST':
			if request.POST.get('submit') == 'class-info':
				class_info_date_str		= request.POST['class-info-date']
				class_info_section		= request.POST['class-info-section']
				
				class_info_date = datetime.strptime(class_info_date_str, '%Y-%m-%d').date()
				class_info_day = class_info_date.strftime("%A")

				calendar = Calendar.objects.filter(date=class_info_date)

				# If calendar instance for that day is not created
				if not calendar.exists():
					context1 = {
						#dash-main
						'no_class_found' : "bilkul_nhi",
						'selected_date' : class_info_date,
						'selected_schedule' : class_info_section,
					}
					context.update(context1)

					return render(request, 'home/dashboard.html', context)

				calendar = calendar[0]

				# If No Class is Scheduled on that day
				if calendar.class_scheduled is False:
					context1 = {
						#dash-main
						'no_class_scheduled' : "haan",
						'calendar_remark' : calendar.remark,
						'selected_date' : class_info_date,
						'selected_schedule' : class_info_section,
					}
					context.update(context1)

					return render(request, 'home/dashboard.html', context)

				schedule = Schedule.objects.filter(day=class_info_day, section=class_info_section)
				if schedule.exists():
					if class_info_date > date.today():
						students_attended = "Class not yet scheduled!"
						vol_volunteered = "Class not yet scheduled!"
					else:
						students_attended = StudentAttendence.objects.filter(date = calendar, present = True).order_by('sid__school_class')
						vol_volunteered = VolunteerAttendence.objects.filter(date = calendar, present = True).order_by('roll_no__roll_no')

						# USE A LIST INSTEAD [G, A, C, M, S]
						# Students present from each village
						stu_att_g = 0
						stu_att_a = 0
						stu_att_c = 0
						stu_att_m = 0
						stu_att_s = 0
						stu_att_ms = 0
						if students_attended.exists():
							for stu_att in students_attended:
								stu_vill = stu_att.sid.village
								if stu_vill == 'G':
									stu_att_g += 1
								elif stu_vill == 'A':
									stu_att_a += 1
									stu_att_ms += 1
								elif stu_vill == 'C':
									stu_att_c += 1
									stu_att_ms += 1
								elif stu_vill == 'M':
									stu_att_m += 1
									stu_att_ms += 1
								elif stu_vill == 'S':
									stu_att_s += 1
									stu_att_ms += 1

						stu_att_village = {
							'stu_att_g': stu_att_g,
							'stu_att_a': stu_att_a,
							'stu_att_c': stu_att_c,
							'stu_att_m': stu_att_m,
							'stu_att_s': stu_att_s,
							'stu_att_ms': stu_att_ms,
						}

						if class_info_date == date.today() and not students_attended.exists():
							students_attended = "Not yet updated!"
						elif not students_attended.exists():
							students_attended = "No student present."

						if class_info_date == date.today() and not vol_volunteered.exists():
							vol_volunteered = "Not yet updated!"
						elif not vol_volunteered.exists():
							vol_volunteered = "No volunteers were present."

					cw_hw = ClassworkHomework.objects.filter(date=class_info_date, section=schedule[0])
					if cw_hw.exists():
						context1 = {
							#dash-main
							'schedule' : schedule,
							'students_attended' : students_attended,
							'stu_att_village' : stu_att_village,
							'vol_volunteered' : vol_volunteered,
							'cw_hw' : cw_hw[0],
							'selected_date' : class_info_date,
							'selected_schedule' : class_info_section,
						}
						context.update(context1)

						return render(request, 'home/dashboard.html', context)
					else:
						context1 = {
							#dash-main
							'schedule' : schedule,
							'students_attended' : students_attended,
							'stu_att_village' : stu_att_village,
							'vol_volunteered' : vol_volunteered,
							'cw_hw' : {
								'cw' : "Not yet updated!",
								'hw' : "Not yet updated!",
							},
							'selected_date' : class_info_date,
							'selected_schedule' : class_info_section,
						}
						context.update(context1)

						return render(request, 'home/dashboard.html', context)
				else:
					# Keeping it just is case if Ajax takes time to load. DISABLE SUBMIT BUTTON UNTIL AJAX LOADS.
					context1 = {
						#dash-main
						'no_schedule_found' : "yup!",
						'selected_date' : class_info_date,
						'selected_schedule' : class_info_section,
					}
					context.update(context1)

					return render(request, 'home/dashboard.html', context)  # The chosen section is not taught on the chosen day

			# elif request.POST.get('submit') == 'cwhw-date':
			# 	cwhw_date_str = request.POST['date']

			# 	cwhw_date = datetime.strptime(cwhw_date_str, '%Y-%m-%d').date()
			# 	cwhw_day = cwhw_date.strftime("%A")

			# 	calendar_date = Calendar.objects.filter(date = cwhw_date)

			# 	if calendar_date.exists():
			# 		context1 = {
			# 			#dash-main
			# 			'class_info_submitted' : "nopes",

			# 			#dash-cwhw
			# 			'cwhw_selected_date' : cwhw_date,
			# 			'cwhw_section': Schedule.objects.filter(day=cwhw_day),
			# 		}
			# 		context.update(context1)

			# 		return render(request, 'home/dashboard.html', context)
			# 	else:
			# 		context1 = {
			# 			#dash-main
			# 			'class_info_submitted' : "nopes",

			# 			#dash-cwhw
			# 			'cwhw_selected_date' : cwhw_date,
			# 			'cwhw_error' : "The chosen day is not yet updated in the Calender.",
			# 		}
			# 		context.update(context1)

			# 		return render(request, 'home/dashboard.html', context)


			# TRY UPDATING CW_HW WITH AJAX
			elif request.POST.get('submit') == 'update-cwhw':
				# Date and section selected before pressing submit button
				cwhw_date_str			= request.POST['date']
				cwhw_section			= request.POST['section']
				cw						= request.POST['cw']
				hw						= request.POST['hw']
				comment					= request.POST['comment']

				cwhw_date = datetime.strptime(cwhw_date_str, '%Y-%m-%d').date()
				cwhw_day = cwhw_date.strftime("%A")

				# cwhw_selected_date = datetime.strptime(cwhw_selected_date_str, '%Y-%m-%d').date()

				# if cwhw_selected_date == cwhw_date:
				cal_date = Calendar.objects.get(date = cwhw_date)
				sch_section = Schedule.objects.get(day=cwhw_day, section=cwhw_section)

				if ClassworkHomework.objects.filter(date=cal_date, section=sch_section).exists():
					cw_hw = ClassworkHomework.objects.get(date=cal_date, section=sch_section)
					if cw:
						cw_hw.cw += '\n' + cw + '\n - ' + volun.first_name + ' ' + volun.last_name + ', ' + volun.roll_no + '\n'
					if hw:
						cw_hw.hw += '\n' + hw + '\n - ' + volun.first_name + ' ' + volun.last_name + ', ' + volun.roll_no + '\n'
					if comment:
						cw_hw.comment += '\n' + comment + '\n - ' + volun.first_name + ' ' + volun.last_name + ', ' + volun.roll_no + '\n'
					cw_hw.save()
				else:
					if cw:
						cw += '\n - ' + volun.first_name + ' ' + volun.last_name + ', ' + volun.roll_no + '\n'
					if hw:
						hw += '\n - ' + volun.first_name + ' ' + volun.last_name + ', ' + volun.roll_no + '\n'
					if comment:
						comment += '\n - ' + volun.first_name + ' ' + volun.last_name + ', ' + volun.roll_no + '\n'
					cw_hw = ClassworkHomework(date=cal_date, section=sch_section, cw=cw, hw=hw, comment = comment)
					cw_hw.save()

				context1 = {
					#dash-main
					'class_info_submitted' : "nopes",

					# CAN USE SESSION VARIABLES TO SEND THESE DATA
					'selected_date' : cwhw_date,   # <-- New
					'selected_schedule' : cwhw_section,   # <-- New

					#dash-cwhw
					'toast' : "CW_HW update successful!",
				}
				context.update(context1)

				messages.success(request, 'CW_HW update successful!')
				return render(request, 'home/dashboard.html', context)
				# else:
				# 	context1 = {
				# 		#dash-main
				# 		'class_info_submitted' : "nopes",

				# 		#dash-cwhw
				# 		'cwhw_selected_date' : cwhw_date,
				# 		'cwhw_error' : "You've changed the date! Kindly submit the chosen date before updating.",
				# 		'toast' : "CW_HW update failed!",
				# 	}
				# 	context.update(context1)

				# 	return render(request, 'home/dashboard.html', context)

		else:
			context1 = {
				#dash-main
				'class_info_submitted' : "nopes",
			}
			context.update(context1)

			return render(request, 'home/dashboard.html', context)
	return redirect('home')

def showSectionInDashboard(request):
	class_info_date_str = request.GET.get('class_date', None)
	class_info_date = datetime.strptime(class_info_date_str, '%Y-%m-%d').date()
	class_info_day = class_info_date.strftime("%A")

	schedule = Schedule.objects.filter(day = class_info_day).order_by('section').values()

	data = {}

	sch_list = []
	for sch in schedule:
		sch_list.append(sch)

	data['schedule'] = sch_list
	data['choices'] = {key: value for key, value in Schedule.SECTION}
	
	# json_data = json.dumps(data)
	# json_schedule = serializers.serialize("json", schedule)
	return JsonResponse(data)

@login_required
def updateSchedule(request):
	volun = Volunteer.objects.filter(email=request.user)
	if volun.exists():
		volun = volun[0]
	else:
		return redirect('set_profile')

	context = {
		#dash-schedule
		'day': Schedule.DAY,
		'section': Schedule.SECTION,
		'update_req' : UpdateScheduleRequest.objects.filter(volunteer = volun).order_by('-date'),
		'last_update_req' : UpdateScheduleRequest.objects.filter(volunteer = volun, approved = False, declined = False, by_admin = False, cancelled = False),
	}
	
	if request.method == 'POST':
		if request.POST.get('submit') == 'update-schedule':
			day			= request.POST['day']
			section		= request.POST['section']

			schedule = Schedule.objects.get(day=day, section=section)

			prev_update_req = UpdateScheduleRequest.objects.filter(volunteer=volun, approved=False, declined=False, by_admin=False, cancelled=False)
			if prev_update_req.exists():
				prev_update_req = prev_update_req[0]
				prev_update_req.cancelled = True
				# For cpanel
				# prev_update_req.approved = False
				# prev_update_req.declined = False
				# prev_update_req.by_admin = False
				prev_update_req.save()

			prev_vol_sch = VolunteerSchedule.objects.filter(roll_no = volun)
			if prev_vol_sch.exists():
				prev_sch = prev_vol_sch[0].schedule
				update_req = UpdateScheduleRequest(volunteer = volun, previous_schedule = prev_sch, updated_schedule = schedule)
				update_req.save()
			else:
				update_req = UpdateScheduleRequest(volunteer = volun, updated_schedule = schedule)
				update_req.save()

			messages.success(request, 'Schedule update requested successfully!')
			return HttpResponseRedirect(reverse('update_schedule'))

		elif request.POST.get('submit') == 'cancel-last-req':
			prev_update_req = UpdateScheduleRequest.objects.filter(volunteer=volun, approved=False, declined=False, by_admin=False, cancelled=False)
			if prev_update_req.exists():
				prev_update_req = prev_update_req[0]
				prev_update_req.cancelled = True
				# For cpanel
				# prev_update_req.approved = False
				# prev_update_req.declined = False
				# prev_update_req.by_admin = False
				prev_update_req.save()

			messages.success(request, 'Last request cancelled successfully!')
			return HttpResponseRedirect(reverse('update_schedule'))

	return render(request, 'home/update_schedule.html', context)

def showSectionInUpdateSchedule(request):
	sch_day = request.GET.get('sch_day', None)

	schedule = Schedule.objects.filter(day = sch_day).order_by('section').values()

	data = {}

	sch_list = []
	for sch in schedule:
		sch_list.append(sch)

	data['schedule'] = sch_list
	data['choices'] = {key: value for key, value in Schedule.SECTION}
	
	# json_data = json.dumps(data)
	# json_schedule = serializers.serialize("json", schedule)
	return JsonResponse(data)

@login_required
def updateProfile(request):
	volun = Volunteer.objects.filter(email=request.user)
	if volun.exists():
		volun = volun[0]
	else:
		return redirect('set_profile')

	context = {
		#dash-update
		'last_4_year': datetime.now().year - 4,
	}
	
	if request.method == 'POST':
		roll_no         = request.POST['roll_no']
		first_name      = request.POST['first_name']
		last_name       = request.POST['last_name']
		gender          = request.POST['gender']
		alt_email       = request.POST['alt_email']
		batch           = request.POST['batch']
		programme       = request.POST['programme']
		street_address1 = request.POST['street_address1']
		street_address2 = request.POST['street_address2']
		pincode         = request.POST['pincode']
		city            = request.POST['city']
		state           = request.POST['state']
		dob             = request.POST['dob']
		contact_no      = request.POST['contact_no']

		if roll_no:
			if volun.roll_no != roll_no:
				duplicate_roll_check = Volunteer.objects.filter(roll_no=roll_no)
				if duplicate_roll_check.exists():
					context1 = {
						'update_error': "A volunteer with entered roll no. already exists."
					}
					context.update(context1)
					messages.error(request, 'Profile update failed!')
					return render(request, 'home/update_profile.html', context)
				else:
					volun.roll_no = roll_no
		if first_name:
			volun.first_name = first_name
		if last_name:
			volun.last_name = last_name
		volun.gender = gender
		volun.batch = batch
		volun.programme = programme
		volun.dob = dob
		if contact_no:
			volun.contact_no = contact_no
		if alt_email:
			volun.alt_email = alt_email
		if street_address1:
			volun.street_address1 = street_address1
		if street_address2:
			volun.street_address2 = street_address2
		if city:
			volun.city = city
		if state:
			volun.state = state
		if pincode:
			volun.pincode = pincode

		volun.save()

		messages.success(request, 'Profile updated Successfully!')
		return HttpResponseRedirect(reverse('update_profile'))

	return render(request, 'home/update_profile.html', context)

@login_required
def studentsAttendence(request):
	volun = Volunteer.objects.filter(email=request.user)
	if volun.exists():
		volun = volun[0]
	else:
		return redirect('set_profile')

	today_cal = Calendar.objects.filter(date=date.today())
	# Update today's date in Calendar if not already there
	if today_cal.exists():
		today_cal = today_cal[0]
	else:
		today_cal_new = Calendar(date = date.today())
		today_cal_new.save()
		today_cal = Calendar.objects.get(date=date.today())

	# For Creating Empty Student Attendence Instances
	if today_cal.class_scheduled is True and not StudentAttendence.objects.filter(date__date = date.today()).exists():
		today_stu_sch = StudentSchedule.objects.filter(day=date.today().strftime("%A"))
		for stu_sch in today_stu_sch:
			stu_attendance = StudentAttendence(sid = stu_sch.sid, date = today_cal)
			stu_attendance.save()

	if today_cal.class_scheduled is True and StudentAttendence.objects.filter(date__date = date.today()).count() != StudentSchedule.objects.filter(day=date.today().strftime("%A")).count():
		today_stu_sch = StudentSchedule.objects.filter(day=date.today().strftime("%A"))
		for stu_sch in today_stu_sch:
			if not StudentAttendence.objects.filter(sid = stu_sch.sid, date = today_cal).exists():
				stu_attendance = StudentAttendence(sid = stu_sch.sid, date = today_cal)
				stu_attendance.save()

	# If no Class is Scheduled
	no_class_today = ''
	if today_cal.class_scheduled is False:
		no_class_today = 'yesss!'

	context = {
		# Overall
		'no_class_today' : no_class_today,

		#dash-stu-att
		'today_date' : date.today(),
		# 'today_stu' : StudentAttendence.objects.filter(date = today_cal).order_by('sid__school_class'),
		'today_stu' : StudentAttendence.objects.filter(date = today_cal, sid__school_class__range = (1, 3)).order_by('sid__school_class'),
	}

	if request.method == 'POST':
		today_date = Calendar.objects.get(date=date.today())
		stu_array = request.POST.getlist('attended')
		selected_class = request.POST['selected_class']

		class_range = selected_class.split('-')
		class_range_min = class_range[0]
		class_range_max = class_range[1]

		# Mark everyone's absent
		stu_today = StudentAttendence.objects.filter(date = today_date, sid__school_class__range = (class_range_min, class_range_max))
		for stu in stu_today:
			stu.present = False
			stu.hw_done = False
			stu.save()

		for sid in stu_array:
			stu = Student.objects.get(id=sid)
			stu_attendance = StudentAttendence.objects.filter(sid = stu, date = today_date)[0]
			stu_attendance.present = True
			stu_attendance.hw_done = False
			stu_attendance.save()

		messages.success(request, 'Attendence marked successfully!')
		return HttpResponseRedirect(reverse('stu_attendence'))

	return render(request, 'home/stu_attendence.html', context)

def studentAttendenceAjax(request):
	stu_class = request.GET.get('stu_class', None)

	today_date = Calendar.objects.get(date=date.today())

	class_range = stu_class.split('-')
	class_range_min = class_range[0]
	class_range_max = class_range[1]

	stu_to_show = StudentAttendence.objects.filter(date = today_date, sid__school_class__range = (class_range_min, class_range_max)).order_by('sid__school_class').values()
	
	data = {}

	stu_list = []
	for stu in stu_to_show:
		stu_list.append(stu)

	students = Student.objects.all()
	stu_dict = {}
	
	for stu in students:
		stu_dict[stu.id] = model_to_dict(stu)

	data['stu_today'] = stu_list
	data['students'] = stu_dict
	
	
	# json_data = json.dumps(data)
	# json_stu_to_show = serializers.serialize("json", stu_to_show)
	return JsonResponse(data)

@login_required
def volunteersAttendence(request):
	volun = Volunteer.objects.filter(email=request.user)
	if volun.exists():
		volun = volun[0]
	else:
		return redirect('set_profile')

	today_cal = Calendar.objects.filter(date=date.today())
	# Update today's date in Calendar if not already there
	if today_cal.exists():
		today_cal = today_cal[0]
	else:
		today_cal_new = Calendar(date = date.today())
		today_cal_new.save()
		today_cal = Calendar.objects.get(date=date.today())

	# For Creating Empty Volunteer Attendence Instances
	if today_cal.class_scheduled is True and not VolunteerAttendence.objects.filter(date__date = date.today()).exists():
		today_vol_sch = VolunteerSchedule.objects.filter(day=date.today().strftime("%A"))
		for vol_sch in today_vol_sch:
			vol_attendance = VolunteerAttendence(roll_no = vol_sch.roll_no, date = today_cal)
			vol_attendance.save()
	
	# If no Class is Scheduled
	no_class_today = ''
	if not today_cal.class_scheduled:
		no_class_today = 'yesss!'

	context = {
		# Overall
		'no_class_today' : no_class_today,

		#dash-vol-att
		'today_date' : date.today(),
		'today_volun' : VolunteerAttendence.objects.filter(date = today_cal).order_by('roll_no__roll_no'),
	}
	
	if request.method == 'POST':
		today_date = Calendar.objects.get(date=date.today())
		vol_array = request.POST.getlist('volunteered')
		extra_vol_array = request.POST.getlist('extra-vol')

		# Mark everyone's absent
		vol_today = VolunteerAttendence.objects.filter(date = today_date)
		for vol in vol_today:
			vol.present = False

			# For cpanel
			if vol.extra is None:
				vol.extra = False

			vol.save()

		for vol in vol_array:
			roll_no = Volunteer.objects.get(id=vol) #i is first column of volun_array
			vol_attendance = VolunteerAttendence.objects.get(roll_no = roll_no, date = today_date)
			vol_attendance.present = True

			# For cpanel
			if vol_attendance.extra is None:
				vol_attendance.extra = False

			vol_attendance.save()

		for extra_vol in extra_vol_array:
			roll_no = Volunteer.objects.filter(roll_no = extra_vol)
			if roll_no.exists():
				extra_vol_att = VolunteerAttendence.objects.filter(roll_no = roll_no[0], date = today_date)
				if not extra_vol_att.exists():
					extra_vol_att = VolunteerAttendence(roll_no = roll_no[0], date = today_date, present = True, extra = True)
					extra_vol_att.save()

		messages.success(request, 'Attendence marked successfully!')
		return HttpResponseRedirect(reverse('vol_attendence'))

	return render(request, 'home/vol_attendence.html', context)

@login_required
def volunteersList(request):
	context = {
		'day': Schedule.DAY,
	}
	return render(request, 'home/vol_list.html', context)

def volunteerListAjax(request):
	vol_list_day = request.GET.get('vol_list_day', None)

	vol_to_show = VolunteerSchedule.objects.filter(day = vol_list_day).order_by('schedule__section').values()

	data = {}

	vol_list = []
	for vol in vol_to_show:
		vol_list.append(vol)

	volunteers = Volunteer.objects.all()
	vol_dict = {}
	
	for vol in volunteers:
		vol_dict[vol.id] = model_to_dict(vol)

	schedules = Schedule.objects.all()
	sch_dict = {}
	
	for sch in schedules:
		sch_dict[sch.id] = model_to_dict(sch)

	data['stu_today'] = vol_list
	data['volunteers'] = vol_dict
	data['schedule'] = sch_dict
	data['choices'] = {key: value for key, value in Schedule.SECTION}
	
	
	# json_data = json.dumps(data, cls=DjangoJSONEncoder)
	# json_vol_list = serializers.serialize("json", vol_list)
	return JsonResponse(data)

def feedback(request):
	submitted = ''
	if request.method == 'POST':
		anonymous_array	= request.POST.getlist('anonymousCheck')
		name			= request.POST['name']
		roll_no			= request.POST['rollNo']
		email			= request.POST['email']
		feedback		= request.POST['feedback']
		if len(anonymous_array) != 0:
			feedback = Feedback(name = 'Anonymous', feedback = feedback)
			feedback.save()
		else:
			feedback = Feedback(name=name, roll_no=roll_no, email=email, feedback=feedback)
			feedback.save()

		submitted = "yesssss!"
	context = {
		'logout_redirect_site' : 'feedback',
		'submitted' : submitted,
	}
	return render(request, 'home/feedback.html', context)

def update_students(request):
	path = "D:/Python Projects/JagratiWebApp/Jagrati/home/student.xlsx"

	wb_obj = load_workbook(path)

	sheet_obj = wb_obj.active 

	max_row = sheet_obj.max_row

	for i in range(3, max_row+1):
		first_name = sheet_obj.cell(row = i, column = 2).value
		last_name = sheet_obj.cell(row = i, column = 3).value
		school_class = sheet_obj.cell(row = i, column = 4).value
		village = sheet_obj.cell(row = i, column = 5).value
		guardian_name = sheet_obj.cell(row = i, column = 6).value
		contact_no = sheet_obj.cell(row = i, column = 7).value

		if not Student.objects.filter(first_name = first_name, last_name = last_name, school_class = school_class, village = village, guardian_name = guardian_name).exists():
			if contact_no is None:
				student = Student(first_name = first_name, last_name = last_name, school_class = school_class, village = village, guardian_name = guardian_name)
				student.save()
			else:
				student = Student(first_name = first_name, last_name = last_name, school_class = school_class, village = village, guardian_name = guardian_name, contact_no = int(contact_no))
				student.save()
			for day, day2 in Schedule.DAY:
				print(day)
				stu_sch = StudentSchedule(sid = student, schedule = Schedule.objects.get(day = day, section = '4A'))
				stu_sch.save()
	
	return HttpResponse('Updated successfully!')