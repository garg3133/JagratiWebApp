from datetime import date, datetime, timedelta

import os

#Django
from django.db.models import Count, Q
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import (
    login_required, user_passes_test, permission_required
)
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy

# third-party
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from home.models import Calendar, Schedule
from home.views import has_authenticated_profile
from .models import Student, StudentAttendance, StudentSchedule
from .forms import StudentModelForm

# GLOBAL VARIABLES
today_date = date.today()
today_day = today_date.strftime("%w")


# VIEWS FUNCTIONS

@login_required
@user_passes_test(
    has_authenticated_profile,
    login_url=reverse_lazy('accounts:complete_profile')
)
@permission_required('students.delete_student', raise_exception=True)
def index(request):
    students = Student.objects.order_by('school_class', 'first_name', 'last_name')
    context = {
        "students": students
    }
    return render(request, 'students/index.html', context)


@login_required
@user_passes_test(
    has_authenticated_profile,
    login_url=reverse_lazy('accounts:complete_profile')
)
def profile(request, pk):
    """View student profile."""
    stu_profile = get_object_or_404(Student, id=pk)
    stu_schedule = StudentSchedule.objects.filter(
        student=stu_profile).order_by('day')
    context = {
        'profile': stu_profile,
        'stu_schedule': stu_schedule,
    }
    return render(request, 'students/profile.html', context)


@login_required
@user_passes_test(
    has_authenticated_profile,
    login_url=reverse_lazy('accounts:complete_profile')
)
# @permission_required('students.add_student', raise_exception=True)
def add_student(request):
    """Add new student."""
    if request.method == 'POST':
        form = StudentModelForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save()

            schedules_9th = Schedule.objects.filter(section__section_id='4')
            stu_schedules = []
            for schedule in schedules_9th:
                stu_schedule = StudentSchedule(student=student, day=schedule.day, schedule=schedule)
                stu_schedules.append(stu_schedule)
            StudentSchedule.objects.bulk_create(stu_schedules)

            messages.success(request, "Student added successfully!")
        else:
            # Todo: Send form data and errors back to page.
            messages.error(request, "Something went wrong. Please try again!")
        return redirect('students:add_student')

    context = {
        'villages': Student.VILLAGE,
    }
    return render(request, 'students/add_student.html', context)


@login_required
@user_passes_test(
    has_authenticated_profile,
    login_url=reverse_lazy('accounts:complete_profile')
)
# @permission_required('students.change_student', raise_exception=True)
def update_profile(request, pk):
    """Update student profile."""
    profile = get_object_or_404(Student, id=pk)
    villages = Student.VILLAGE

    context = {
        'profile': profile,
        'villages': villages,
        'form': StudentModelForm(instance=profile),
    }

    if 'profile_image' in request.FILES:
        # TODO: Update profile picture using AJAX
        profile.profile_image.delete(False)
        profile.profile_image = request.FILES["profile_image"]
        profile.save()

        messages.success(request, "Profile picture updated.")
        context["form"] = StudentModelForm(request.POST, instance=profile)
        return render(request, 'students/update_profile.html', context)

    if request.method == 'POST':
        form = StudentModelForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('students:profile', pk=pk)
        else:
            context["form"] = form
            # TODO: Form errors missing
            messages.error(request, "Something went wrong. Please try again!")

    return render(request, 'students/update_profile.html', context)


@login_required
@user_passes_test(
    has_authenticated_profile,
    login_url=reverse_lazy('accounts:complete_profile')
)
# @permission_required('students.change_student', raise_exception=True)
def verify_profile(request, pk, verify):
    """Mark student profile as verified/unverified."""
    profile = get_object_or_404(Student, id=pk)

    profile.verified = True if verify == 1 else False
    profile.save()

    return redirect('students:profile', pk=pk)


@login_required
@user_passes_test(
    has_authenticated_profile,
    login_url=reverse_lazy('accounts:complete_profile')
)
# @permissions_required
def attendance(request):
    """Student attendance page."""
    today_cal = Calendar.objects.filter(date=today_date)
    # TO BE REMOVED...
    # Update today's date in Calendar if not already there
    if today_cal.exists():
        today_cal = today_cal[0]
    else:
        today_cal_new = Calendar(date=today_date)
        today_cal_new.save()
        today_cal = Calendar.objects.get(date=today_date)
    # ...TILL HERE

    context = {
        'today_date': today_date,
    }

    if not today_cal.class_scheduled:
        context['no_class_today'] = True
        return render(request, 'students/attendance.html', context)

    today_stu_sch = StudentSchedule.objects.select_related(
        'student').filter(day=today_day)
    today_stu_att = StudentAttendance.objects.filter(cal_date__date=today_date)

    if not today_stu_att.exists():
        # Create Empty Student Attendance Instances
        stu_att = []
        for stu_sch in today_stu_sch:
            stu_att.append(StudentAttendance(
                student=stu_sch.student, cal_date=today_cal))
        StudentAttendance.objects.bulk_create(stu_att)
    elif today_stu_att.count() != today_stu_sch.count():
        # Some new students added in today's schedule (not necessarily present today)
        stu_att = []
        for stu_sch in today_stu_sch:
            if not today_stu_att.filter(student=stu_sch.student).exists():
                stu_att.append(StudentAttendance(
                    student=stu_sch.student, cal_date=today_cal))
        StudentAttendance.objects.bulk_create(stu_att)

    context['stu_att_today'] = StudentAttendance.objects.filter(
        cal_date=today_cal, student__school_class__range=(1, 3)).order_by(
        'student__school_class', 'student__first_name', 'student__last_name')

    return render(request, 'students/attendance.html', context)


@login_required
@user_passes_test(
    has_authenticated_profile, redirect_field_name=None,
    login_url=reverse_lazy('accounts:complete_profile')
)
# @permissions_required
def ajax_fetch_students(request):
    """Fetch students based on class-group selected on student attendance page."""
    today_cal = Calendar.objects.get(date=today_date)
    data = {}
    stu_class = request.GET.get('stu_class', None)
    class_range_min, class_range_max = stu_class.split('-')

    stu_att_today = StudentAttendance.objects.filter(
        cal_date=today_cal, student__school_class__range=(
            class_range_min, class_range_max),
    ).order_by('student__school_class', 'student__first_name', 'student__last_name')

    for stu_att in stu_att_today:
        # key --> For sorting purpose.
        key = str(stu_att.student.school_class) + stu_att.student.get_full_name
        data[key] = [stu_att.id, stu_att.student.id, stu_att.student.get_verified_name,
                     stu_att.student.school_class, stu_att.present]

    return JsonResponse(data)


@login_required
@user_passes_test(
    has_authenticated_profile, redirect_field_name=None,
    login_url=reverse_lazy('accounts:complete_profile')
)
# @permissions_required
def ajax_mark_attendance(request):
    """Mark/unmark student attendance."""
    stu_id = request.GET['std_id']
    is_present = request.GET['is_present']
    stu_att = StudentAttendance.objects.get(
        student__id=stu_id, cal_date=today_date)
    stu_att.present = True if is_present == 'true' else False
    stu_att.save()
    data = {'success': True}
    return JsonResponse(data)


@login_required
@user_passes_test(
    has_authenticated_profile, redirect_field_name=None,
    login_url=reverse_lazy('accounts:complete_profile')
)
# @permissions_required
def ajax_mark_homework(request):
    """Mark/unmark homework done."""
    if request.method == 'POST' and request.is_ajax():
        stu_id = request.POST.get('stu_id')
        homework_done = request.POST.get('homework_done')

        date_str = request.POST.get('date')
        date = datetime.strptime(date_str, "%Y-%m-%d").date()
        cal_date = Calendar.objects.get(date=date)

        student = StudentAttendance.objects.get(
            student__id=stu_id, cal_date=cal_date)
        student.hw_done = True if homework_done == 'true' else False
        student.save()
        data = {'success': True}
        return JsonResponse(data)


@login_required
@user_passes_test(
    has_authenticated_profile,
    login_url=reverse_lazy('accounts:complete_profile')
)
@permission_required('students.add_student', raise_exception=True)
def update_from_sheets(request):
    """Create Student model instances from excel sheet."""
    if request.method == 'POST':
        sheet = request.FILES['sheet']
        # Temporarily save the file
        fs = FileSystemStorage(location=settings.TEMP_ROOT)
        filename = fs.save(sheet.name, sheet)
        file_path = os.path.join(settings.TEMP_ROOT, filename)
        # Process the file
        wb_obj = load_workbook(file_path)
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row

        for i in range(2, max_row + 1):
            id = sheet_obj.cell(row=i, column=1).value
            first_name = sheet_obj.cell(row=i, column=2).value
            last_name = sheet_obj.cell(row=i, column=3).value
            gender = sheet_obj.cell(row=i, column=4).value
            school_class = sheet_obj.cell(row=i, column=5).value
            village = sheet_obj.cell(row=i, column=6).value
            guardian_name = sheet_obj.cell(row=i, column=7).value
            contact_no = str(sheet_obj.cell(row=i, column=8).value or '')
            remarks = sheet_obj.cell(row=i, column=9).value or ''

            # If id is already present, don't do anything (any changes to details should only be done using the portal)
            # If id is not present, check if that student is already in the database (to avoid duplicates)
            if (id is None and
                not Student.objects.filter(first_name=first_name, last_name=last_name, school_class=school_class,
                                           village=village, guardian_name=guardian_name).exists()):
                student = Student(first_name=first_name, last_name=last_name, gender=gender, school_class=school_class,
                                  village=village, guardian_name=guardian_name, contact_no=contact_no, remarks=remarks)
                student.save()

                schedules_9th = Schedule.objects.filter(section__section_id='4')
                stu_schedules = []
                for schedule in schedules_9th:
                    stu_schedule = StudentSchedule(student=student, day=schedule.day, schedule=schedule)
                    stu_schedules.append(stu_schedule)
                StudentSchedule.objects.bulk_create(stu_schedules)

        # Delete the file
        os.remove(file_path)
        messages.success(request, "Data Updated Successfully")
        return redirect('students:index')

    return render(request, 'students/update_from_sheets.html')


@login_required
@user_passes_test(
    has_authenticated_profile,
    login_url=reverse_lazy('accounts:complete_profile')
)
@permission_required('students.view_student', raise_exception=True)
def generate_sheet(request):
    students = Student.objects.all().order_by('school_class', 'first_name', 'last_name')

    file_path = os.path.join(settings.MEDIA_ROOT, 'students', 'sheets','student_profile_data.xlsx')
    excel = Workbook()
    sheet = excel.active

    col_width = [8, 20, 20, 8, 8, 8, 30, 20, 30, 8]

    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'first_name'
    sheet.cell(row=1, column=3).value = 'last_name'
    sheet.cell(row=1, column=4).value = 'gender'
    sheet.cell(row=1, column=5).value = 'school_class'
    sheet.cell(row=1, column=6).value = 'village'
    sheet.cell(row=1, column=7).value = 'guardian_name'
    sheet.cell(row=1, column=8).value = 'contact_no'
    sheet.cell(row=1, column=9).value = 'remarks'
    sheet.cell(row=1, column=10).value = 'verified'

    for i in range(1, 11):
        sheet.cell(row=1, column=i).font = Font(size=12, bold=True)
        sheet.column_dimensions[chr(65+(i-1))].width=col_width[i-1]

    for row, student in enumerate(students, 2): 
        sheet.cell(row=row, column=1).value = student.id 
        sheet.cell(row=row, column=2).value = student.first_name
        sheet.cell(row=row, column=3).value = student.last_name
        sheet.cell(row=row, column=4).value = student.gender
        sheet.cell(row=row, column=5).value = student.school_class
        sheet.cell(row=row, column=6).value = student.village
        sheet.cell(row=row, column=7).value = student.guardian_name
        sheet.cell(row=row, column=8).value = student.contact_no
        sheet.cell(row=row, column=9).value = student.remarks
        sheet.cell(row=row, column=10).value = student.verified

    excel.save(file_path)

    return HttpResponseRedirect(settings.MEDIA_URL + '/students/sheets/student_profile_data.xlsx')

@login_required
@user_passes_test(
    has_authenticated_profile, redirect_field_name=None,
    login_url=reverse_lazy('accounts:complete_profile')
)
def leaderboard(request):
    date_from = date.today() - timedelta(days=30)
    date_to = date.today()

    students = StudentAttendance.objects.filter(cal_date__date__range=[date_from, date_to], present=True)
    students = students.values('student_id', 'student__first_name', 'student__last_name',
                               'student__school_class', 'student__village').annotate(
                                   total_attendance=Count('cal_date'),
                                   total_hw_done=Count('cal_date', filter=Q(hw_done=True))
                               ).order_by('-total_attendance', '-total_hw_done', 'student__school_class', 'student__first_name')
                                    
    return render(request, 'students/leaderboard.html', {'students': students})
